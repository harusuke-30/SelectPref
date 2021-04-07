#! python3
# coding: utf-8

import bs4
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.drawing.image import Image
import os
import requests

wb = openpyxl.Workbook()
ws = wb.active
w_row = 2

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'
}

def get_retry(url, retry_times, errs, headers):
    for t in range(retry_times + 1):
        r =requests.get(url, headers=headers)
        if t < retry_times:
            if r.status_code in errs:
                time.sleep(2)
                continue
        return r

def getProducts():

    ws.cell(row=1,column=1).value = 'チェック'
    ws.cell(row=1,column=2).value = 'img'
    ws.cell(row=1,column=3).value = '商品'
    ws.cell(row=1,column=4).value = '発送元'
    ws.cell(row=1,column=5).value = '価格'
    ws.cell(row=1,column=6).value = 'URL'

    w_row = 2
    url1 = 'https://www.mercari.com/jp/search/?page='
    url2 =  '&keyword=%E3%83%AB%E3%83%BC%E3%83%95%E3%83%9C%E3%83%83%E3%82%AF%E3%82%B9&sort_order=&category_root=1318&category_child=115&category_grand_child%5B1115%5D=1&brand_name=&brand_id=&size_group=&price_min=&price_max=&status_on_sale=1&_s=U2FsdGVkX185KkRQR-2qUe7J3Gp15qOIdJ658cPkx9Y8qoCSdWlPwHv9abPpZ-t8dFIIXNazo9bLmPSd0kQgD6kdLepNuxANGoXpxERRnL4'

    flg = 0
    for i in range(1,5):
        url = url1 + str(i) + url2
#        print(url)
        res = get_retry(url, 5, [501, 502, 503],headers)
        if res.status_code in [501, 502, 503]:
            return []
        res.raise_for_status()
        # 検索結果のリンクを取得
        soup = bs4.BeautifulSoup(res.text, 'html5lib')

#        for products in soup.select('.items-box'):
        for a in soup.select('.items-box a'):
#            print('####',a.get('href'))
            url2 = 'https://www.mercari.com' + a.get('href')

#            print('##',url2)
            res2 = get_retry(url2, 5, [501, 502, 503],headers)
            if res2.status_code in [501, 502, 503]:
                print(res2.status_code)
                return []
            res2.raise_for_status()
            # 検索結果のリンクを取得
            soup2 = bs4.BeautifulSoup(res2.text, 'html5lib')

#            print('#',str(soup2).encode('cp932','ignore'))
#            print('##')
            for a,b in zip(soup2.select('.item-detail-table th'), soup2.select('.item-detail-table td')):
#                print('########',b)

                if( a.getText() == '配送元地域'):
#                    print('#',a.getText(),'##',b.getText())
                    if( b.getText() == '富山県' or b.getText() == '石川県'):
                        print('#',a.getText(),'##',b.getText())
                        print('##',url2)
#                print('#',b.getText())
#                print(b.select('td').string)

#                print(b.select('tr')[0].gettext())

#            print('###',str(products.select('.items-box')).encode('cp932',"ignore"))
#            url = products.select('a')[0].get('href')
#            url = products.select('.items-box a')[0].get('href')
#            print(url)
#            flg = 1
#            break
        if (flg == 1):
            break

getProducts()
